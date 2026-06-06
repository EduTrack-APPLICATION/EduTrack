"""
Servicio de generación de reportes PDF (ReportLab) y Excel (OpenPyXL).
"""
import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Estudiante, Grupo, Evaluacion, Nota
from app.services.calculo_service import CalculoService


# Paleta de colores corporativos
COLOR_PRIMARY = colors.HexColor('#3b82f6')
COLOR_SECONDARY = colors.HexColor('#6366f1')
COLOR_SUCCESS = colors.HexColor('#10b981')
COLOR_DANGER = colors.HexColor('#ef4444')
COLOR_WARNING = colors.HexColor('#f59e0b')
COLOR_DARK = colors.HexColor('#1e293b')
COLOR_LIGHT = colors.HexColor('#f1f5f9')

# Paleta específica para diplomas — tonos medalla
COLOR_GOLD       = colors.HexColor('#c9a227')   # oro mate
COLOR_GOLD_LIGHT = colors.HexColor('#f4cf66')
COLOR_GOLD_DARK  = colors.HexColor('#8e7218')
COLOR_SILVER     = colors.HexColor('#94a3b8')
COLOR_SILVER_LT  = colors.HexColor('#cbd5e1')
COLOR_SILVER_DK  = colors.HexColor('#475569')
COLOR_BRONZE     = colors.HexColor('#a16207')
COLOR_BRONZE_LT  = colors.HexColor('#d97706')
COLOR_BRONZE_DK  = colors.HexColor('#713f12')
COLOR_NAVY       = colors.HexColor('#1e3a8a')
COLOR_NAVY_LT    = colors.HexColor('#3b82f6')
COLOR_PAPER      = colors.HexColor('#fdfbf5')   # crema muy claro (papel)
COLOR_INK        = colors.HexColor('#2c2416')   # tinta marrón oscura


class ReporteService:
    """Generación de reportes en PDF y Excel."""

    # ---------- PDF ----------

    @staticmethod
    def _estilo_base():
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='TituloReporte',
            fontSize=20, leading=24, alignment=TA_CENTER,
            textColor=COLOR_DARK, fontName='Helvetica-Bold',
            spaceAfter=10,
        ))
        styles.add(ParagraphStyle(
            name='Subtitulo',
            fontSize=12, leading=16, alignment=TA_CENTER,
            textColor=colors.grey, spaceAfter=20,
        ))
        styles.add(ParagraphStyle(
            name='SeccionTitulo',
            fontSize=14, leading=18,
            textColor=COLOR_PRIMARY, fontName='Helvetica-Bold',
            spaceAfter=8, spaceBefore=12,
        ))
        return styles

    @staticmethod
    def _encabezado(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(COLOR_PRIMARY)
        canvas.rect(0, A4[1] - 1.5 * cm, A4[0], 1.5 * cm, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 14)
        canvas.drawString(2 * cm, A4[1] - 1 * cm, 'EduTrack')
        canvas.setFont('Helvetica', 9)
        canvas.drawRightString(A4[0] - 2 * cm, A4[1] - 1 * cm,
                                f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        # Pie
        canvas.setFillColor(colors.grey)
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(A4[0] / 2, 1 * cm,
                                  f'Página {doc.page} - EduTrack © Sistema de Gestión Académica')
        canvas.restoreState()

    @staticmethod
    def boletin_estudiante(estudiante, grupo=None):
        """Genera boletín PDF de un estudiante (todos sus grupos o uno específico)."""
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=2.5 * cm, bottomMargin=2 * cm,
            leftMargin=2 * cm, rightMargin=2 * cm
        )
        styles = ReporteService._estilo_base()
        elements = []

        # Título
        elements.append(Paragraph('BOLETÍN ACADÉMICO', styles['TituloReporte']))
        elements.append(Paragraph(f'Estudiante: {estudiante.nombre_completo}', styles['Subtitulo']))

        # Datos del estudiante
        datos = [
            ['Código:', estudiante.codigo, 'Cédula:', estudiante.cedula or 'N/A'],
            ['Email:', estudiante.email or 'N/A', 'Teléfono:', estudiante.telefono or 'N/A'],
            ['Sección:', estudiante.seccion or 'N/A', 'Estado:', estudiante.estado.upper()],
        ]
        tbl_datos = Table(datos, colWidths=[3 * cm, 5 * cm, 3 * cm, 5 * cm])
        tbl_datos.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, -1), COLOR_LIGHT),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [COLOR_LIGHT, colors.white]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tbl_datos)
        elements.append(Spacer(1, 0.6 * cm))

        grupos = [grupo] if grupo else list(estudiante.grupos)
        if not grupos:
            elements.append(Paragraph('El estudiante no está matriculado en ningún grupo.',
                                       styles['Normal']))
        else:
            for g in grupos:
                elements.append(Paragraph(
                    f'{g.materia.nombre} — {g.codigo}', styles['SeccionTitulo']))
                elements.append(Paragraph(
                    f'<b>Profesor:</b> {g.profesor_titular.nombre_completo} | '
                    f'<b>Periodo:</b> {g.periodo}',
                    styles['Normal']))
                elements.append(Spacer(1, 0.3 * cm))

                # Tabla de notas
                data = [['Evaluación', 'Tipo', 'Fecha', 'Puntaje', 'Máximo', '%', 'Peso']]
                evaluaciones = Evaluacion.query.filter_by(grupo_id=g.id).order_by(
                    Evaluacion.fecha).all()
                for ev in evaluaciones:
                    nota = Nota.query.filter_by(
                        estudiante_id=estudiante.id, evaluacion_id=ev.id).first()
                    if nota:
                        pct = (nota.puntaje / ev.puntaje_maximo * 100) if ev.puntaje_maximo else 0
                        data.append([
                            ev.nombre[:30],
                            ev.tipo.capitalize(),
                            ev.fecha.strftime('%d/%m/%Y'),
                            f'{nota.puntaje:.1f}',
                            f'{ev.puntaje_maximo:.0f}',
                            f'{pct:.1f}%',
                            f'{ev.porcentaje:.0f}%',
                        ])
                    else:
                        data.append([
                            ev.nombre[:30], ev.tipo.capitalize(),
                            ev.fecha.strftime('%d/%m/%Y'),
                            '—', f'{ev.puntaje_maximo:.0f}', '—', f'{ev.porcentaje:.0f}%',
                        ])

                tabla = Table(data, colWidths=[5 * cm, 2.2 * cm, 2.2 * cm, 1.8 * cm,
                                                1.8 * cm, 1.5 * cm, 1.5 * cm])
                tabla.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('PADDING', (0, 0), (-1, -1), 5),
                ]))
                elements.append(tabla)
                elements.append(Spacer(1, 0.3 * cm))

                # Resumen y estado
                calc = CalculoService.nota_final_estudiante(estudiante.id, g.id)
                asist = CalculoService.porcentaje_asistencia(estudiante.id, g.id)
                estado_key, estado_label, _ = CalculoService.estado_academico(
                    calc['nota_final'], calc['porcentaje_acumulado'])

                color_estado = COLOR_SUCCESS if 'aprob' in estado_key else (
                    COLOR_WARNING if 'recup' in estado_key or 'riesgo' in estado_key else COLOR_DANGER)

                resumen_data = [
                    ['Nota Final', 'Acumulado', 'Asistencia', 'Estado'],
                    [
                        f'{calc["nota_final"]:.2f}',
                        f'{calc["porcentaje_acumulado"]:.0f}%',
                        f'{asist:.1f}%' if asist is not None else 'N/A',
                        estado_label,
                    ]
                ]
                tbl_resumen = Table(resumen_data, colWidths=[4 * cm, 4 * cm, 4 * cm, 4 * cm])
                tbl_resumen.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), COLOR_DARK),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('BACKGROUND', (3, 1), (3, 1), color_estado),
                    ('TEXTCOLOR', (3, 1), (3, 1), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 11),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('PADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
                ]))
                elements.append(tbl_resumen)
                elements.append(Spacer(1, 0.8 * cm))

        doc.build(elements, onFirstPage=ReporteService._encabezado,
                  onLaterPages=ReporteService._encabezado)
        buf.seek(0)
        return buf

    @staticmethod
    def reporte_grupo_pdf(grupo):
        """Reporte PDF de todos los estudiantes de un grupo con sus notas finales."""
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            topMargin=2.5 * cm, bottomMargin=2 * cm,
            leftMargin=1.5 * cm, rightMargin=1.5 * cm
        )
        styles = ReporteService._estilo_base()
        elements = []

        elements.append(Paragraph(f'REPORTE DE GRUPO — {grupo.codigo}', styles['TituloReporte']))
        elements.append(Paragraph(
            f'{grupo.materia.nombre} | Profesor: {grupo.profesor_titular.nombre_completo} | '
            f'Periodo: {grupo.periodo}',
            styles['Subtitulo']))

        # Tabla con todas las evaluaciones como columnas
        evaluaciones = Evaluacion.query.filter_by(grupo_id=grupo.id).order_by(
            Evaluacion.fecha).all()

        header = ['#', 'Estudiante']
        for ev in evaluaciones:
            header.append(f'{ev.nombre[:15]}\n({ev.porcentaje:.0f}%)')
        header += ['Nota Final', 'Estado']

        data = [header]
        estudiantes = list(grupo.estudiantes.order_by('apellido'))
        for i, est in enumerate(estudiantes, 1):
            row = [str(i), est.nombre_completo[:30]]
            for ev in evaluaciones:
                nota = Nota.query.filter_by(
                    estudiante_id=est.id, evaluacion_id=ev.id).first()
                row.append(f'{nota.puntaje:.1f}' if nota else '—')

            calc = CalculoService.nota_final_estudiante(est.id, grupo.id)
            _, estado_label, _ = CalculoService.estado_academico(
                calc['nota_final'], calc['porcentaje_acumulado'])
            row.append(f'{calc["nota_final"]:.2f}')
            row.append(estado_label)
            data.append(row)

        col_widths = [0.8 * cm, 5 * cm] + [2 * cm] * len(evaluaciones) + [2 * cm, 2.5 * cm]
        tabla = Table(data, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(tabla)

        doc.build(elements, onFirstPage=ReporteService._encabezado,
                  onLaterPages=ReporteService._encabezado)
        buf.seek(0)
        return buf

    # ---------- EXCEL ----------

    @staticmethod
    def _aplicar_estilo_header(ws, row, cols):
        """Aplica estilo de encabezado a una fila."""
        font_header = Font(bold=True, color='FFFFFF', size=11)
        fill_header = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='cccccc'),
            right=Side(style='thin', color='cccccc'),
            top=Side(style='thin', color='cccccc'),
            bottom=Side(style='thin', color='cccccc'),
        )
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align
            cell.border = border

    @staticmethod
    def notas_grupo_excel(grupo):
        """Excel con notas de todos los estudiantes de un grupo."""
        wb = Workbook()
        ws = wb.active
        ws.title = f'Notas {grupo.codigo}'[:31]

        # Encabezado superior
        ws['A1'] = 'EduTrack — Reporte de Notas'
        ws['A1'].font = Font(bold=True, size=14, color='1e293b')
        ws.merge_cells('A1:F1')

        ws['A2'] = f'Grupo: {grupo.codigo} — {grupo.materia.nombre}'
        ws['A3'] = f'Profesor: {grupo.profesor_titular.nombre_completo}'
        ws['A4'] = f'Periodo: {grupo.periodo} | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        for r in [2, 3, 4]:
            ws[f'A{r}'].font = Font(size=10, color='64748b')

        # Headers
        start_row = 6
        headers = ['#', 'Código', 'Estudiante']
        evaluaciones = Evaluacion.query.filter_by(grupo_id=grupo.id).order_by(
            Evaluacion.fecha).all()
        for ev in evaluaciones:
            headers.append(f'{ev.nombre} ({ev.porcentaje:.0f}%)')
        headers += ['Nota Final', 'Acumulado %', 'Asistencia %', 'Estado']

        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=h)
        ReporteService._aplicar_estilo_header(ws, start_row, len(headers))

        # Datos
        estudiantes = list(grupo.estudiantes.order_by('apellido'))
        border_thin = Border(
            left=Side(style='thin', color='e2e8f0'),
            right=Side(style='thin', color='e2e8f0'),
            top=Side(style='thin', color='e2e8f0'),
            bottom=Side(style='thin', color='e2e8f0'),
        )

        for i, est in enumerate(estudiantes, 1):
            row = start_row + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=est.codigo)
            ws.cell(row=row, column=3, value=est.nombre_completo)

            col = 4
            for ev in evaluaciones:
                nota = Nota.query.filter_by(
                    estudiante_id=est.id, evaluacion_id=ev.id).first()
                ws.cell(row=row, column=col, value=nota.puntaje if nota else None)
                col += 1

            calc = CalculoService.nota_final_estudiante(est.id, grupo.id)
            asist = CalculoService.porcentaje_asistencia(est.id, grupo.id)
            _, estado_label, color = CalculoService.estado_academico(
                calc['nota_final'], calc['porcentaje_acumulado'])

            ws.cell(row=row, column=col, value=round(calc['nota_final'], 2))
            ws.cell(row=row, column=col + 1, value=calc['porcentaje_acumulado'])
            ws.cell(row=row, column=col + 2, value=asist if asist is not None else 'N/A')
            estado_cell = ws.cell(row=row, column=col + 3, value=estado_label)

            # Color por estado
            color_map = {
                'success': 'd1fae5',
                'warning': 'fef3c7',
                'danger': 'fee2e2',
                'info': 'dbeafe',
            }
            if color in color_map:
                estado_cell.fill = PatternFill(
                    start_color=color_map[color], end_color=color_map[color], fill_type='solid')

            # Bordes y alineación
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = border_thin
                if c >= 4:
                    cell.alignment = Alignment(horizontal='center')

        # Ancho de columnas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        for col in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.freeze_panes = f'D{start_row + 1}'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def asistencia_grupo_excel(grupo):
        """Excel con asistencia de un grupo, una columna por fecha."""
        wb = Workbook()
        ws = wb.active
        ws.title = f'Asistencia {grupo.codigo}'[:31]

        ws['A1'] = f'EduTrack — Asistencia: {grupo.codigo}'
        ws['A1'].font = Font(bold=True, size=14)

        from app.models import Asistencia
        fechas = sorted({a.fecha for a in grupo.asistencias.all()})

        headers = ['#', 'Estudiante'] + [f.strftime('%d/%m') for f in fechas] + ['% Asist.']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        ReporteService._aplicar_estilo_header(ws, 3, len(headers))

        estudiantes = list(grupo.estudiantes.order_by('apellido'))
        simbolos = {'presente': 'P', 'ausente': 'A', 'justificado': 'J', 'tardia': 'T'}

        for i, est in enumerate(estudiantes, 1):
            ws.cell(row=3 + i, column=1, value=i)
            ws.cell(row=3 + i, column=2, value=est.nombre_completo)
            col = 3
            for f in fechas:
                asist = Asistencia.query.filter_by(
                    estudiante_id=est.id, grupo_id=grupo.id, fecha=f).first()
                ws.cell(row=3 + i, column=col,
                        value=simbolos.get(asist.estado, '-') if asist else '-')
                col += 1
            pct = CalculoService.porcentaje_asistencia(est.id, grupo.id)
            ws.cell(row=3 + i, column=col, value=f'{pct}%' if pct is not None else 'N/A')

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 8

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def _paleta_diploma(posicion):
        """Retorna la paleta de colores según la posición (1=oro, 2=plata, 3=bronce, otros=azul)."""
        if posicion == 1:
            return {'main': COLOR_GOLD, 'light': COLOR_GOLD_LIGHT, 'dark': COLOR_GOLD_DARK,
                    'label': 'PRIMER LUGAR', 'medal_text': '1'}
        elif posicion == 2:
            return {'main': COLOR_SILVER, 'light': COLOR_SILVER_LT, 'dark': COLOR_SILVER_DK,
                    'label': 'SEGUNDO LUGAR', 'medal_text': '2'}
        elif posicion == 3:
            return {'main': COLOR_BRONZE, 'light': COLOR_BRONZE_LT, 'dark': COLOR_BRONZE_DK,
                    'label': 'TERCER LUGAR', 'medal_text': '3'}
        elif isinstance(posicion, int):
            return {'main': COLOR_NAVY, 'light': COLOR_NAVY_LT, 'dark': COLOR_DARK,
                    'label': f'{posicion}º LUGAR', 'medal_text': str(posicion)}
        else:
            return {'main': COLOR_NAVY, 'light': COLOR_NAVY_LT, 'dark': COLOR_DARK,
                    'label': 'RECONOCIMIENTO', 'medal_text': '★'}

    @staticmethod
    def _dibujar_decoracion_diploma(canvas, paleta):
        """Dibuja todos los elementos decorativos del diploma sobre el canvas."""
        from reportlab.lib.pagesizes import landscape, A4
        w, h = landscape(A4)
        c = canvas
        c.saveState()

        # Fondo color papel crema
        c.setFillColor(COLOR_PAPER)
        c.rect(0, 0, w, h, fill=1, stroke=0)

        # === MARCO ORNAMENTAL ===
        # Marco exterior grueso (color medalla)
        c.setStrokeColor(paleta['main'])
        c.setLineWidth(4)
        c.rect(0.7 * cm, 0.7 * cm, w - 1.4 * cm, h - 1.4 * cm, stroke=1, fill=0)

        # Línea ornamental intermedia (más delgada)
        c.setStrokeColor(paleta['light'])
        c.setLineWidth(0.8)
        c.rect(0.95 * cm, 0.95 * cm, w - 1.9 * cm, h - 1.9 * cm, stroke=1, fill=0)

        # Marco interior fino
        c.setStrokeColor(paleta['dark'])
        c.setLineWidth(0.5)
        c.rect(1.15 * cm, 1.15 * cm, w - 2.3 * cm, h - 2.3 * cm, stroke=1, fill=0)

        # === FLOURISHES EN LAS ESQUINAS ===
        # Pequeñas curvas decorativas en cada esquina
        def _flourish(cx, cy, angle):
            c.saveState()
            c.translate(cx, cy)
            c.rotate(angle)
            c.setStrokeColor(paleta['main'])
            c.setLineWidth(1.2)
            # Pequeña curva tipo "rama"
            p = c.beginPath()
            p.moveTo(0, 0)
            p.curveTo(0.8 * cm, 0.05 * cm, 1.4 * cm, 0.4 * cm, 1.6 * cm, 1 * cm)
            c.drawPath(p, stroke=1, fill=0)
            # Hojita
            p2 = c.beginPath()
            p2.moveTo(0.6 * cm, 0.15 * cm)
            p2.curveTo(0.5 * cm, 0.5 * cm, 0.85 * cm, 0.85 * cm, 1.05 * cm, 0.65 * cm)
            p2.curveTo(0.95 * cm, 0.35 * cm, 0.75 * cm, 0.2 * cm, 0.6 * cm, 0.15 * cm)
            c.setFillColor(paleta['light'])
            c.drawPath(p2, stroke=1, fill=1)
            # Punto decorativo
            c.setFillColor(paleta['main'])
            c.circle(0.15 * cm, 0.15 * cm, 0.08 * cm, fill=1, stroke=0)
            c.restoreState()

        # 4 esquinas (ajustar ángulos)
        _flourish(1.4 * cm, 1.4 * cm, 0)
        _flourish(w - 1.4 * cm, 1.4 * cm, 90)
        _flourish(w - 1.4 * cm, h - 1.4 * cm, 180)
        _flourish(1.4 * cm, h - 1.4 * cm, 270)

        # === MEDALLA / SELLO CIRCULAR EN LA PARTE SUPERIOR ===
        medal_cx = w / 2
        medal_cy = h - 2.4 * cm
        medal_r = 1.1 * cm

        # Sombra de la medalla
        c.setFillColor(colors.Color(0, 0, 0, alpha=0.12))
        c.circle(medal_cx + 0.06 * cm, medal_cy - 0.06 * cm, medal_r, fill=1, stroke=0)

        # Círculo exterior de la medalla
        c.setFillColor(paleta['main'])
        c.circle(medal_cx, medal_cy, medal_r, fill=1, stroke=0)

        # Anillo más claro
        c.setFillColor(paleta['light'])
        c.circle(medal_cx, medal_cy, medal_r * 0.85, fill=1, stroke=0)

        # Anillo interior (color principal de nuevo)
        c.setFillColor(paleta['main'])
        c.circle(medal_cx, medal_cy, medal_r * 0.68, fill=1, stroke=0)

        # Centro brillante
        c.setFillColor(paleta['light'])
        c.circle(medal_cx, medal_cy, medal_r * 0.55, fill=1, stroke=0)

        # === RAYITAS DE LA MEDALLA (estilo sol) — pequeños puntos decorativos ===
        c.setFillColor(paleta['dark'])
        import math
        for angle_deg in range(0, 360, 30):
            angle = math.radians(angle_deg)
            x = medal_cx + math.cos(angle) * medal_r * 0.6
            y = medal_cy + math.sin(angle) * medal_r * 0.6
            c.circle(x, y, 0.04 * cm, fill=1, stroke=0)

        # === NÚMERO / SÍMBOLO EN LA MEDALLA ===
        c.setFillColor(paleta['dark'])
        c.setFont('Helvetica-Bold', 24)
        c.drawCentredString(medal_cx, medal_cy - 0.28 * cm, paleta['medal_text'])

        # === CINTAS DE LA MEDALLA ===
        # Lazo izquierdo
        c.setFillColor(paleta['main'])
        p1 = c.beginPath()
        p1.moveTo(medal_cx - medal_r * 0.7, medal_cy - medal_r * 0.7)
        p1.lineTo(medal_cx - medal_r * 1.1, medal_cy - medal_r * 1.8)
        p1.lineTo(medal_cx - medal_r * 0.5, medal_cy - medal_r * 1.5)
        p1.lineTo(medal_cx - medal_r * 0.35, medal_cy - medal_r * 0.95)
        p1.close()
        c.drawPath(p1, stroke=0, fill=1)
        # Lazo derecho
        p2 = c.beginPath()
        p2.moveTo(medal_cx + medal_r * 0.7, medal_cy - medal_r * 0.7)
        p2.lineTo(medal_cx + medal_r * 1.1, medal_cy - medal_r * 1.8)
        p2.lineTo(medal_cx + medal_r * 0.5, medal_cy - medal_r * 1.5)
        p2.lineTo(medal_cx + medal_r * 0.35, medal_cy - medal_r * 0.95)
        p2.close()
        c.drawPath(p2, stroke=0, fill=1)
        # Hendiduras de las cintas (color oscuro)
        c.setFillColor(paleta['dark'])
        p3 = c.beginPath()
        p3.moveTo(medal_cx - medal_r * 0.8, medal_cy - medal_r * 1.45)
        p3.lineTo(medal_cx - medal_r * 0.5, medal_cy - medal_r * 1.5)
        p3.lineTo(medal_cx - medal_r * 0.65, medal_cy - medal_r * 1.7)
        p3.close()
        c.drawPath(p3, stroke=0, fill=1)
        p4 = c.beginPath()
        p4.moveTo(medal_cx + medal_r * 0.8, medal_cy - medal_r * 1.45)
        p4.lineTo(medal_cx + medal_r * 0.5, medal_cy - medal_r * 1.5)
        p4.lineTo(medal_cx + medal_r * 0.65, medal_cy - medal_r * 1.7)
        p4.close()
        c.drawPath(p4, stroke=0, fill=1)

        c.restoreState()

    @staticmethod
    def _dibujar_pie_diploma(canvas, paleta, periodo_txt=''):
        """Dibuja el pie del diploma con línea de firma y sello de la institución."""
        from reportlab.lib.pagesizes import landscape, A4
        w, h = landscape(A4)
        c = canvas
        c.saveState()

        # === LÍNEA DE FIRMA (izquierda) ===
        firma_x = 4 * cm
        firma_y = 2.6 * cm
        c.setStrokeColor(COLOR_INK)
        c.setLineWidth(0.6)
        c.line(firma_x, firma_y, firma_x + 5.5 * cm, firma_y)
        c.setFillColor(COLOR_INK)
        c.setFont('Helvetica-Oblique', 9)
        c.drawCentredString(firma_x + 2.75 * cm, firma_y - 0.4 * cm,
                            'Director / Coordinador Académico')

        # === SELLO INSTITUCIONAL (derecha) ===
        sello_cx = w - 6.5 * cm
        sello_cy = 3.2 * cm
        sello_r = 1.35 * cm
        # Marco circular doble del sello
        c.setStrokeColor(paleta['dark'])
        c.setLineWidth(1.2)
        c.circle(sello_cx, sello_cy, sello_r, fill=0, stroke=1)
        c.setLineWidth(0.5)
        c.circle(sello_cx, sello_cy, sello_r * 0.92, fill=0, stroke=1)
        # Pequeños puntitos decorativos alrededor del anillo
        import math as _math
        c.setFillColor(paleta['main'])
        for ang_deg in range(0, 360, 30):
            ang = _math.radians(ang_deg)
            px = sello_cx + _math.cos(ang) * sello_r * 0.96
            py = sello_cy + _math.sin(ang) * sello_r * 0.96
            c.circle(px, py, 0.05 * cm, fill=1, stroke=0)
        # Texto arriba: "EDUTRACK"
        c.setFillColor(paleta['dark'])
        c.setFont('Helvetica-Bold', 9)
        c.drawCentredString(sello_cx, sello_cy + sello_r * 0.45, 'E D U T R A C K')
        # Estrella central
        c.setFillColor(paleta['main'])
        c.setFont('Helvetica-Bold', 22)
        c.drawCentredString(sello_cx, sello_cy - sello_r * 0.18, '★')
        # Texto debajo de la estrella
        c.setFillColor(paleta['dark'])
        c.setFont('Helvetica-Bold', 6)
        c.drawCentredString(sello_cx, sello_cy - sello_r * 0.55, 'CUADRO DE HONOR')
        # Texto pequeño abajo
        if periodo_txt:
            c.setFillColor(COLOR_INK)
            c.setFont('Helvetica-Oblique', 8)
            c.drawCentredString(sello_cx, sello_cy - sello_r - 0.4 * cm, periodo_txt)

        c.restoreState()

    @staticmethod
    def _construir_elementos_diploma(item, periodo=None, anio=None):
        """Construye la lista de flowables (texto) para un diploma."""
        styles = getSampleStyleSheet()
        paleta = ReporteService._paleta_diploma(item.get('posicion'))

        # Estilos refinados con jerarquía clara
        estilo_lugar = ParagraphStyle(
            'LugarDip', parent=styles['Normal'],
            fontSize=10, textColor=paleta['dark'],
            alignment=TA_CENTER, fontName='Helvetica-Bold',
            spaceAfter=2,
        )
        estilo_titulo = ParagraphStyle(
            'TituloDip', parent=styles['Title'],
            fontSize=40, textColor=COLOR_INK,
            alignment=TA_CENTER, fontName='Times-Bold',
            spaceAfter=2, leading=44,
        )
        estilo_subtitulo = ParagraphStyle(
            'SubDip', parent=styles['Normal'],
            fontSize=10, textColor=paleta['dark'],
            alignment=TA_CENTER, fontName='Helvetica-Oblique',
            spaceAfter=14, leading=12,
        )
        estilo_texto = ParagraphStyle(
            'TxtDip', parent=styles['Normal'],
            fontSize=11, textColor=COLOR_INK,
            alignment=TA_CENTER, fontName='Times-Italic',
            leading=14, spaceAfter=4,
        )
        estilo_nombre = ParagraphStyle(
            'NomDip', parent=styles['Title'],
            fontSize=28, textColor=paleta['main'],
            alignment=TA_CENTER, fontName='Times-Bold',
            spaceAfter=2, leading=32,
        )
        estilo_separador = ParagraphStyle(
            'SepDip', parent=styles['Normal'],
            fontSize=8, textColor=paleta['main'],
            alignment=TA_CENTER, fontName='Helvetica-Bold',
            spaceAfter=4,
        )
        estilo_codigo = ParagraphStyle(
            'CodDip', parent=styles['Normal'],
            fontSize=8, textColor=paleta['dark'],
            alignment=TA_CENTER, fontName='Helvetica',
            spaceAfter=10,
        )
        estilo_promedio_label = ParagraphStyle(
            'PromLab', parent=styles['Normal'],
            fontSize=8, textColor=paleta['dark'],
            alignment=TA_CENTER, fontName='Helvetica-Bold',
            spaceAfter=1,
        )
        estilo_promedio = ParagraphStyle(
            'PromDip', parent=styles['Title'],
            fontSize=26, textColor=paleta['main'],
            alignment=TA_CENTER, fontName='Times-Bold',
            spaceAfter=2, leading=28,
        )
        estilo_materias = ParagraphStyle(
            'MatDip', parent=styles['Normal'],
            fontSize=8, textColor=paleta['dark'],
            alignment=TA_CENTER, fontName='Helvetica-Oblique',
        )

        estudiante = item['estudiante']
        promedio = item['promedio']
        grupos_eval = item['grupos_evaluados']

        elementos = []

        # Espacio para que el contenido empiece debajo de la medalla
        elementos.append(Spacer(1, 2.5 * cm))

        # Título principal grande
        elementos.append(Paragraph('Diploma de Honor', estilo_titulo))

        # Subtítulo decorativo
        elementos.append(Paragraph(
            '— por excelencia académica —', estilo_subtitulo
        ))

        # "PRIMER LUGAR" en su propio lugar después del título (no detrás de la medalla)
        lugar_text = '   '.join(list(paleta['label'].replace(' ', '·')))
        elementos.append(Paragraph(
            f'<font color="#{paleta["main"].hexval()[2:]}">{lugar_text}</font>',
            estilo_lugar
        ))

        elementos.append(Spacer(1, 0.2 * cm))

        # Texto introductorio
        elementos.append(Paragraph(
            'Se otorga el presente reconocimiento a', estilo_texto
        ))

        elementos.append(Spacer(1, 0.05 * cm))

        # Nombre del estudiante en grande
        elementos.append(Paragraph(estudiante.nombre_completo, estilo_nombre))

        elementos.append(Spacer(1, 0.15 * cm))

        if estudiante.codigo:
            elementos.append(Paragraph(
                f'Código de estudiante: <b>{estudiante.codigo}</b>',
                estilo_codigo
            ))
        else:
            elementos.append(Spacer(1, 0.2 * cm))

        # Texto del cuerpo
        elementos.append(Paragraph(
            'por su destacado desempeño académico, demostrando dedicación,<br/>'
            'esfuerzo y excelencia durante sus estudios.',
            estilo_texto
        ))

        elementos.append(Spacer(1, 0.25 * cm))

        # Promedio: label arriba + número grande
        elementos.append(Paragraph('PROMEDIO  GENERAL', estilo_promedio_label))
        elementos.append(Paragraph(f'{promedio:.2f}', estilo_promedio))

        if grupos_eval:
            elementos.append(Paragraph(
                f'sobre {grupos_eval} '
                f'{"materia evaluada" if grupos_eval == 1 else "materias evaluadas"}',
                estilo_materias
            ))

        return elementos, paleta

    @staticmethod
    def diploma_honor(item, periodo=None, anio=None):
        """
        Genera un diploma PDF de cuadro de honor para un estudiante.

        `item` es un dict con la forma producida por CalculoService.top_estudiantes():
            { 'posicion': int, 'estudiante': Estudiante, 'promedio': float,
              'grupos_evaluados': int, 'grupos': [(Grupo, nota_final)] }

        Retorna un BytesIO con el PDF en horizontal (landscape).
        """
        from reportlab.lib.pagesizes import landscape, A4

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=3 * cm, rightMargin=3 * cm,
            topMargin=1.2 * cm, bottomMargin=3 * cm,
            title='Diploma Cuadro de Honor'
        )

        elementos, paleta = ReporteService._construir_elementos_diploma(
            item, periodo, anio)

        # Texto del periodo para el sello
        periodo_txt = ''
        if periodo and anio:
            periodo_txt = f'Periodo {periodo} · {anio}'
        elif anio:
            periodo_txt = f'Año {anio}'
        elif periodo:
            periodo_txt = f'Periodo {periodo}'
        else:
            periodo_txt = f'Emitido el {datetime.now().strftime("%d/%m/%Y")}'

        def _dibujar_pagina(canvas, doc_):
            ReporteService._dibujar_decoracion_diploma(canvas, paleta)
            ReporteService._dibujar_pie_diploma(canvas, paleta, periodo_txt)

        doc.build(elementos, onFirstPage=_dibujar_pagina, onLaterPages=_dibujar_pagina)
        buf.seek(0)
        return buf

    @staticmethod
    def diplomas_lote(items, periodo=None, anio=None):
        """
        Genera un PDF único con varios diplomas (uno por página).
        `items` es una lista de dicts como los retornados por top_estudiantes().
        """
        from reportlab.lib.pagesizes import landscape, A4

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=3 * cm, rightMargin=3 * cm,
            topMargin=1.2 * cm, bottomMargin=3 * cm,
            title='Diplomas Cuadro de Honor'
        )

        # Texto del periodo común
        if periodo and anio:
            periodo_txt = f'Periodo {periodo} · {anio}'
        elif anio:
            periodo_txt = f'Año {anio}'
        elif periodo:
            periodo_txt = f'Periodo {periodo}'
        else:
            periodo_txt = f'Emitido el {datetime.now().strftime("%d/%m/%Y")}'

        # Construir flowables de TODOS los diplomas, separados por PageBreak.
        # Como cada diploma tiene su propia paleta (oro/plata/etc.),
        # guardamos las paletas por número de página y el callback de página
        # las consulta vía un contador mutable.
        paletas_por_pagina = []
        elementos_totales = []

        for idx, item in enumerate(items):
            if idx > 0:
                elementos_totales.append(PageBreak())
            elementos, paleta = ReporteService._construir_elementos_diploma(
                item, periodo, anio)
            elementos_totales.extend(elementos)
            paletas_por_pagina.append(paleta)

        contador = {'pagina': 0}

        def _dibujar(canvas, doc_):
            # `doc_.page` es 1-indexed; le restamos 1 para indexar paletas_por_pagina
            idx = doc_.page - 1
            if idx >= len(paletas_por_pagina):
                idx = len(paletas_por_pagina) - 1
            paleta = paletas_por_pagina[idx]
            ReporteService._dibujar_decoracion_diploma(canvas, paleta)
            ReporteService._dibujar_pie_diploma(canvas, paleta, periodo_txt)

        doc.build(elementos_totales, onFirstPage=_dibujar, onLaterPages=_dibujar)
        buf.seek(0)
        return buf
