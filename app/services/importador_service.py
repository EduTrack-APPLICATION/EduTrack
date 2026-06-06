"""
Servicio de importación masiva de estudiantes desde Excel/CSV.

Flujo:
 1. parsear(archivo)  → lista de dicts con datos crudos
 2. validar(lista)    → lista de validaciones por fila (errores y advertencias)
 3. importar(lista)   → inserción transaccional en BD

Los pasos están separados para permitir vista previa antes de confirmar.
"""
import io
import re
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app import db
from app.models import Estudiante, EstadoEstudianteEnum


# Definición de columnas esperadas y su mapeo a campos del modelo
COLUMNAS = [
    {'key': 'codigo',             'label': 'Código',              'requerido': True,
     'tipo': 'str',   'max_length': 30,
     'ayuda': 'Identificador único del estudiante. Ej: EST-2026001'},
    {'key': 'nombre',             'label': 'Nombre',              'requerido': True,
     'tipo': 'str',   'max_length': 80,
     'ayuda': 'Nombre(s) del estudiante'},
    {'key': 'apellido',           'label': 'Apellido',            'requerido': True,
     'tipo': 'str',   'max_length': 80,
     'ayuda': 'Apellido(s) del estudiante'},
    {'key': 'cedula',             'label': 'Cédula',              'requerido': False,
     'tipo': 'str',   'max_length': 30,
     'ayuda': 'Documento de identidad (opcional pero único si se incluye)'},
    {'key': 'email',              'label': 'Email',               'requerido': False,
     'tipo': 'email', 'max_length': 120,
     'ayuda': 'Correo electrónico (opcional)'},
    {'key': 'telefono',           'label': 'Teléfono',            'requerido': False,
     'tipo': 'str',   'max_length': 30,
     'ayuda': 'Teléfono de contacto'},
    {'key': 'fecha_nacimiento',   'label': 'Fecha de nacimiento', 'requerido': False,
     'tipo': 'date',
     'ayuda': 'Formato: DD/MM/AAAA o AAAA-MM-DD'},
    {'key': 'direccion',          'label': 'Dirección',           'requerido': False,
     'tipo': 'str',   'max_length': 255,
     'ayuda': 'Dirección de residencia'},
    {'key': 'seccion',            'label': 'Sección',             'requerido': False,
     'tipo': 'str',   'max_length': 50,
     'ayuda': 'Sección o curso (ej. 10-A)'},
    {'key': 'encargado',          'label': 'Encargado',           'requerido': False,
     'tipo': 'str',   'max_length': 150,
     'ayuda': 'Nombre del padre, madre o tutor'},
    {'key': 'telefono_encargado', 'label': 'Teléfono encargado',  'requerido': False,
     'tipo': 'str',   'max_length': 30,
     'ayuda': 'Teléfono del padre, madre o tutor'},
]


class ImportadorEstudiantes:
    """Servicio para importación masiva de estudiantes desde Excel/CSV."""

    @staticmethod
    def generar_plantilla():
        """Genera un Excel con encabezados, formato e instrucciones."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Estudiantes'

        headers = [c['label'] + (' *' if c['requerido'] else '') for c in COLUMNAS]
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6',
                                  fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 32

        ejemplos = [
            ['EST-2026100', 'Ana', 'García López', '1-2345-6789',
             'ana.garcia@email.com', '+506 8888-1234', '15/03/2009',
             'San José, Costa Rica', '10-A', 'María López', '+506 8888-9999'],
            ['EST-2026101', 'Carlos', 'Rodríguez Mora', '', '', '',
             '', '', '10-A', '', ''],
        ]

        ejemplo_font = Font(italic=True, color='94A3B8')
        for fila in ejemplos:
            ws.append(fila)
            row_num = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = ejemplo_font
                cell.border = border

        widths = [16, 14, 18, 16, 24, 16, 16, 30, 12, 22, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Hoja de instrucciones
        ws2 = wb.create_sheet('Instrucciones')
        ws2['A1'] = 'Cómo usar esta plantilla'
        ws2['A1'].font = Font(bold=True, size=14, color='1E293B')
        ws2['A2'] = '1. No modifiques los encabezados de la primera hoja.'
        ws2['A3'] = '2. Las columnas con * son obligatorias.'
        ws2['A4'] = '3. Borra las dos filas de ejemplo antes de añadir tus estudiantes reales.'
        ws2['A5'] = '4. El "Código" debe ser único por estudiante.'
        ws2['A6'] = '5. Las fechas pueden ir en formato DD/MM/AAAA o AAAA-MM-DD.'
        ws2['A7'] = '6. Al subir el archivo verás una vista previa antes de confirmar.'
        ws2['A9'] = 'Descripción de los campos:'
        ws2['A9'].font = Font(bold=True, size=12, color='1E293B')

        for i, col in enumerate(COLUMNAS, start=10):
            cell_label = ws2.cell(row=i, column=1, value=f'• {col["label"]}')
            cell_label.font = Font(bold=True)
            ws2.cell(row=i, column=2, value=col['ayuda'])

        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def parsear(file_storage):
        """Lee Excel o CSV y retorna (lista_filas, error_general)."""
        nombre = (file_storage.filename or '').lower()
        if nombre.endswith('.csv'):
            return ImportadorEstudiantes._parsear_csv(file_storage)
        elif nombre.endswith('.xlsx') or nombre.endswith('.xls'):
            return ImportadorEstudiantes._parsear_xlsx(file_storage)
        else:
            return [], 'Formato no soportado. Sube un archivo .xlsx o .csv'

    @staticmethod
    def _parsear_xlsx(file_storage):
        try:
            wb = load_workbook(file_storage, read_only=True, data_only=True)
        except Exception as e:
            return [], f'No se pudo leer el archivo Excel: {e}'

        ws = wb.active
        if ws is None:
            return [], 'El archivo no tiene hojas'

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], 'El archivo no tiene filas de datos'

        header_row = [str(c).strip() if c is not None else '' for c in rows[0]]
        col_to_key = ImportadorEstudiantes._mapear_encabezados(header_row)

        for req in ('codigo', 'nombre', 'apellido'):
            if req not in col_to_key.values():
                return [], f'No se encontró la columna "{req.title()}" en el archivo'

        filas = []
        for fila_idx, row in enumerate(rows[1:], start=2):
            datos = {'_fila': fila_idx}
            tiene_algo = False
            for col_idx, valor in enumerate(row):
                key = col_to_key.get(col_idx)
                if not key:
                    continue
                if valor is None:
                    datos[key] = ''
                elif isinstance(valor, (datetime, date)):
                    datos[key] = valor
                    tiene_algo = True
                else:
                    s = str(valor).strip()
                    datos[key] = s
                    if s:
                        tiene_algo = True
            if tiene_algo:
                filas.append(datos)

        return filas, None

    @staticmethod
    def _parsear_csv(file_storage):
        import csv
        try:
            contenido = file_storage.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file_storage.seek(0)
            try:
                contenido = file_storage.read().decode('latin-1')
            except Exception as e:
                return [], f'No se pudo leer el CSV: {e}'

        sample = contenido[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except csv.Error:
            dialect = csv.excel

        lector = csv.reader(io.StringIO(contenido), dialect=dialect)
        rows = list(lector)
        if len(rows) < 2:
            return [], 'El archivo no tiene filas de datos'

        header_row = [c.strip() for c in rows[0]]
        col_to_key = ImportadorEstudiantes._mapear_encabezados(header_row)

        for req in ('codigo', 'nombre', 'apellido'):
            if req not in col_to_key.values():
                return [], f'No se encontró la columna "{req.title()}" en el archivo'

        filas = []
        for fila_idx, row in enumerate(rows[1:], start=2):
            datos = {'_fila': fila_idx}
            tiene_algo = False
            for col_idx, valor in enumerate(row):
                key = col_to_key.get(col_idx)
                if not key:
                    continue
                s = valor.strip() if valor else ''
                datos[key] = s
                if s:
                    tiene_algo = True
            if tiene_algo:
                filas.append(datos)

        return filas, None

    @staticmethod
    def _mapear_encabezados(header_row):
        aliases = {
            'codigo':             ['codigo', 'código', 'code', 'id'],
            'nombre':             ['nombre', 'nombres', 'name', 'first name', 'firstname'],
            'apellido':           ['apellido', 'apellidos', 'last name', 'lastname', 'surname'],
            'cedula':             ['cedula', 'cédula', 'dni', 'documento', 'identificacion', 'identificación'],
            'email':              ['email', 'correo', 'e-mail', 'mail'],
            'telefono':           ['telefono', 'teléfono', 'phone', 'celular', 'tel'],
            'fecha_nacimiento':   ['fecha de nacimiento', 'fecha nacimiento', 'nacimiento',
                                   'birth date', 'birthday', 'fecha_nacimiento'],
            'direccion':          ['direccion', 'dirección', 'address', 'domicilio'],
            'seccion':            ['seccion', 'sección', 'grado', 'curso', 'class'],
            'encargado':          ['encargado', 'padre', 'madre', 'acudiente', 'tutor', 'guardian'],
            'telefono_encargado': ['telefono encargado', 'teléfono encargado',
                                   'tel encargado', 'tel acudiente', 'phone guardian',
                                   'telefono_encargado'],
        }
        col_to_key = {}
        for idx, header in enumerate(header_row):
            normalizado = header.lower().strip().rstrip('*').strip()
            for key, alts in aliases.items():
                if normalizado in alts or normalizado == key:
                    col_to_key[idx] = key
                    break
        return col_to_key

    @staticmethod
    def validar(filas):
        """Valida cada fila contra reglas del modelo + duplicados."""
        existentes_codigo = {e.codigo for e in
                             Estudiante.query.with_entities(Estudiante.codigo).all()
                             if e.codigo}
        existentes_cedula = {e.cedula for e in
                             Estudiante.query.with_entities(Estudiante.cedula).all()
                             if e.cedula}
        existentes_email  = {e.email.lower() for e in
                             Estudiante.query.with_entities(Estudiante.email).all()
                             if e.email}

        codigos_archivo = {}
        cedulas_archivo = {}
        emails_archivo  = {}

        resultados = []
        for fila in filas:
            errores = []
            advertencias = []
            datos_limpios = {}

            for col in COLUMNAS:
                key = col['key']
                valor = fila.get(key, '')

                if col['requerido'] and not valor:
                    errores.append(f'"{col["label"]}" es obligatorio')
                    continue

                if not valor and not col['requerido']:
                    datos_limpios[key] = None
                    continue

                if col['tipo'] == 'date':
                    fecha = ImportadorEstudiantes._parsear_fecha(valor)
                    if fecha is None:
                        errores.append(f'"{col["label"]}" no es una fecha válida ({valor})')
                    else:
                        datos_limpios[key] = fecha
                elif col['tipo'] == 'email':
                    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', str(valor)):
                        errores.append(f'"{col["label"]}" no parece un email válido')
                    else:
                        datos_limpios[key] = str(valor).strip().lower()
                else:
                    s = str(valor).strip()
                    if col.get('max_length') and len(s) > col['max_length']:
                        errores.append(
                            f'"{col["label"]}" es demasiado largo '
                            f'(máx {col["max_length"]})')
                    else:
                        datos_limpios[key] = s

            # Duplicados
            existe_en_bd = False
            codigo = datos_limpios.get('codigo')
            if codigo:
                if codigo in existentes_codigo:
                    advertencias.append(
                        f'Ya existe un estudiante con código "{codigo}" — se omitirá.')
                    existe_en_bd = True
                elif codigo in codigos_archivo:
                    errores.append(
                        f'Código "{codigo}" repetido en este archivo '
                        f'(también en fila {codigos_archivo[codigo]})')
                else:
                    codigos_archivo[codigo] = fila['_fila']

            cedula = datos_limpios.get('cedula')
            if cedula:
                if cedula in existentes_cedula:
                    advertencias.append(
                        f'Ya existe un estudiante con cédula "{cedula}" — se omitirá.')
                    existe_en_bd = True
                elif cedula in cedulas_archivo:
                    errores.append(
                        f'Cédula "{cedula}" repetida en este archivo '
                        f'(fila {cedulas_archivo[cedula]})')
                else:
                    cedulas_archivo[cedula] = fila['_fila']

            email = datos_limpios.get('email')
            if email:
                if email in existentes_email:
                    advertencias.append(
                        f'Ya existe un estudiante con email "{email}" — se omitirá.')
                    existe_en_bd = True
                elif email in emails_archivo:
                    errores.append(
                        f'Email "{email}" repetido en este archivo '
                        f'(fila {emails_archivo[email]})')
                else:
                    emails_archivo[email] = fila['_fila']

            resultados.append({
                'fila': fila['_fila'],
                'datos': datos_limpios,
                'datos_raw': {k: v for k, v in fila.items() if k != '_fila'},
                'errores': errores,
                'advertencias': advertencias,
                'valido': len(errores) == 0,
                'existe': existe_en_bd,
            })

        return resultados

    @staticmethod
    def _parsear_fecha(valor):
        if isinstance(valor, (datetime, date)):
            return valor.date() if isinstance(valor, datetime) else valor
        s = str(valor).strip()
        formatos = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%m/%d/%Y']
        for fmt in formatos:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def importar(resultados_validados):
        """Importa las filas válidas y no duplicadas. Transaccional."""
        creados = []
        omitidos = 0
        con_error = 0

        for r in resultados_validados:
            if not r['valido']:
                con_error += 1
                continue
            if r['existe']:
                omitidos += 1
                continue

            est = Estudiante(
                codigo=r['datos']['codigo'],
                nombre=r['datos']['nombre'],
                apellido=r['datos']['apellido'],
                cedula=r['datos'].get('cedula'),
                email=r['datos'].get('email'),
                telefono=r['datos'].get('telefono'),
                fecha_nacimiento=r['datos'].get('fecha_nacimiento'),
                direccion=r['datos'].get('direccion'),
                seccion=r['datos'].get('seccion'),
                encargado=r['datos'].get('encargado'),
                telefono_encargado=r['datos'].get('telefono_encargado'),
                estado=EstadoEstudianteEnum.ACTIVO.value,
            )
            db.session.add(est)
            creados.append(est)

        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise

        return {
            'creados': len(creados),
            'omitidos': omitidos,
            'con_error': con_error,
            'estudiantes_creados': creados,
        }
